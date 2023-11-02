import openpyxl
import logging

logger = logging.getLogger(__name__)


def doc2txt(document_path):
    try:
        document_format = document_path.split(".")[-1]
        if document_format == "xlsx":
            wb = openpyxl.load_workbook(document_path)

            text = ''

            for colonne in wb.sheetnames:
                feuille = wb[colonne]

                for ligne in feuille.iter_rows():
                    for cellule in ligne:
                        text += str(cellule.value) + ' '

            wb.close()

            return text

        else:
            raise Exception(
                "Erreur : Merci de fournir un document au format xlsx."
            )

    except Exception as e:
        logger.error("Error : ", str(e))
        return None
